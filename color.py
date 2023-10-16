import openpyxl

# Abre el archivo de Excel
wb = openpyxl.load_workbook('2021.xlsx')
sheet = wb.active

# Itera a través de las filas y columnas para obtener los códigos de color de fondo
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        fill = cell.fill
        if fill.start_color.index:
            color_index = fill.start_color.index
            print(f'Color de fondo en celda {cell.coordinate}: {color_index}')
